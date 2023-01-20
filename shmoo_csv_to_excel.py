#!/usr/bin/python3
#################################################################
#                         shmoo_csv_to_excel                   	#
#################################################################
#                                                               #
#   This script takes in a shmoo CSV and turns the data into    #
#   excel format for viewing                                    #
#                                                               #
#################################################################
# Version 0.2                                                   #
# By Shane Benetz                                               #
# Date: 09.15.2021                                              #
#################################################################
#################################################################
# Version 0.0 is first release 08.06.2021                       #
# Version 0.1 adds graphs, averages, and usage 08.25.2021       #
# Version 0.2 is an update on edge cases for outputDir 09.15.21 #
#################################################################
import argparse
import os
import re
import sys
from copy import copy
import time
from datetime import datetime

from get_timing import get_version
import openpyxl 
from PIL import Image
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter, column_index_from_string
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.chart import BarChart, Series, Reference, LineChart, SurfaceChart
from openpyxl.chart.layout import Layout, ManualLayout

#constants
grey = PatternFill('solid', fgColor='00C0C0C0')
blue = PatternFill('solid', fgColor='DDEBF7') 
green = PatternFill('solid', fgColor='E2EFDA') 
highlight = PatternFill('solid',fgColor='00FFFF00')
orange = PatternFill('solid', fgColor='00FFCC99')
redBg = PatternFill(bgColor='FEB7B7')
thinBorders = border = Border(left=Side(style='thin'),right=Side(style='thin'), \
    top=Side(style='thin'), bottom=Side(style='thin'))
redFont = Font(bold=True, color='00FF0000')

def shmoo_csv_to_excel(csv, outputFile, outputDir, replace, newInfo, commenting, last):
    ''' Control module to write the given CSV file into one worksheet in excel and
    sets up formatting for the worksheet
    Input:
        csv             --> given .csv file to look in
        outputFile      --> name of excel file
        outputDir       --> name of folder to write results to
        replace         --> whether or not to replace current spreadsheet
    Output:
        Saves spreadsheet 
    '''
    # check if input exists
    if not os.path.isfile(csv) :
        return print('%s is not a file' % csv)
    if not csv.endswith('.csv') :
        return print(csv + ' is not a valid .csv file')
    CSVHeader = 'Pattern Name, EQNSET #, Holes, Bad '
    with open(csv,'r') as CSV :
        if not CSVHeader in CSV.readlines()[0]:
            return print(csv + ' is not in the correct format')
    now = datetime.now().strftime('%#m-%#d-%y_%#Hh%Mm%#S') 
    # find correct output location
    name = os.path.basename(os.path.normpath(csv))
    name = name[ : name.find('.csv')]
    if name.count('_') == 2 :
        name = name[:name.find('_')] + '(' +\
            name[name.find('_')+1:name.rfind('_')].replace('p','%') + ')' + \
                 name[name.rfind('_'):]
    if outputFile != None : 
        if os.path.isfile(outputFile) : replace = True
        outputDir = os.path.abspath(os.path.join(outputFile,os.pardir))
    if not (os.path.isdir(outputDir)) :
        os.mkdir(outputDir)
    if not os.access(outputDir, os.W_OK) or not os.access(outputDir, os.R_OK):
            return print('Output directory not accessible')
    if outputFile == None:
        outputFile = name + '_notes.xlsx'
        if os.path.isfile(outputFile) and not replace: 
            outputFile = name + '_notes_' + now + '.xlsx'
    elif not outputFile.endswith('.xlsx') :
        outputFile = outputFile + '.xlsx'
    filename = os.path.join(outputDir,outputFile)
    if os.path.isfile(filename) and not replace :
        outputFile = outputFile.strip('.xlsx') + '_' + now + '.xlsx'
    outputFile = os.path.join(outputDir,os.path.basename(outputFile))
    print('\nWorking on Excel conversion: %s ' % name, end='', flush=True)
    printed = False
    try : # see if file exists
        workbook = load_workbook(filename = outputFile)
        new = False
    except :
        workbook = openpyxl.Workbook()
        new = True
    for worksheet in workbook.sheetnames : 
        if workbook[worksheet].max_column == 1 and workbook[worksheet].max_row == 1:
            del workbook[worksheet] # delete extra empty sheets
        if replace and name == worksheet : # delete sheet to replace
            del workbook[worksheet]
        if worksheet == 'Info' and (replace or newInfo): # delete 
            ws = workbook[worksheet]
            firstRow = 0
            lastRow = ws.max_row
            for r in range(10,ws.max_row):
                firstCell = ws.cell(r,1).value
                if firstRow>0 and firstCell != None and firstCell.endswith('.csv'):
                    lastRow= r-2; break
                if firstCell != None and firstCell == os.path.abspath(csv) :
                    firstRow = r
                if firstRow > 0:
                    for c in range(1,len(ws[r])):
                        ws.cell(r,c).value = None
            if firstRow > 0 :
                ws.row_dimensions.group(firstRow, lastRow, outline_level=0, hidden=False)
                maxRow = ws.max_row
                colsToRemove = lastRow+2-firstRow
                for r in range(lastRow+2,ws.max_row) :
                    ws.row_dimensions[r-colsToRemove] = copy(ws.row_dimensions[r])
                    ws.row_dimensions[r].outlineLevel = 0
                    ws.row_dimensions[r].hidden = False
                if lastRow != maxRow : 
                    ws.delete_rows(firstRow,(lastRow-firstRow)+2)
    if not replace and name in workbook.sheetnames : # make new sheet name
        print('\nCreated new worksheet ' + name + '_' + now + ' in ' + outputFile)
        printed = True
        WS = workbook.create_sheet(name + '_' + now)
    else :
        WS = workbook.create_sheet(name)

    if (not name.replace('Vm','Tm') in workbook.sheetnames) or replace:
        make_info_tab(workbook, csv, last, replace, new) # write paths from CSV 

    # write in the basic headers
    WS['A1'] = 'Updated: ' + datetime.today().ctime(); WS['B2'] = 'TIM#'
    WS['A2'] = 'Pattern Name'; WS['C1'] = 'Issues found in Shmoo'
    WS['C1'].fill = orange
    WS.merge_cells(start_row=1, start_column=3, end_row=2, end_column=5)
    WS['G2'] = 'From'; WS['H2'] = 'To'
    WS.row_dimensions[1].height = 30

    # use helper to write the csv to the worksheet
    try:
        WS = csv_to_worksheet(WS,csv,commenting)
    except Exception as e:
        print('Cannot read ' + csv); print(e); printed = True
        
    # delete empty rows at the end
    for row in WS.rows:
        if row[0].value == None :
            WS.delete_rows(row[0].row)
    #get averages
    startTempCol = 0
    for i in range(10,WS.max_column+6) : # +6 in case added 5 average columns
        if 'C' in str(WS.cell(1, i).value) and startTempCol == 0:
            startTempCol = i -1
        elif not 'C' in str(WS.cell(1, i).value) and startTempCol > 0:
            if startTempCol != i-1:
                WS.insert_cols(i,1)
                WS.cell(1,column=i).value = 'Average'
                WS.cell(2,column=i).value = WS.cell(1,column=i-1).value + ' - '\
                    +  re.sub('[0-9]','',WS.cell(2,column=i-1).value)
                colLetter = get_column_letter(i)
                for row in WS.rows :
                    if row[0].row > 2 :
                        startCol = get_column_letter(startTempCol+1)
                        endCol = get_column_letter(i-1)
                        rowNum = str(row[0].row)
                        # if average is worse than spec
                        if '-' in WS['I1'].value or '-' in WS['F1'].value:
                            rule = CellIsRule(operator='greaterThan', \
                            formula=['$I%s'% rowNum], fill=redBg)
                        else :
                            rule = CellIsRule(operator='lessThan', \
                            formula=['$I%s'% rowNum], fill=redBg)
                        WS.conditional_formatting.add(colLetter+rowNum, rule)
                        row[i-1].value = \
                        '=IFERROR(ROUND(AVERAGE(%s%s:%s%s),2),"FAIL")' % \
                             (startCol,rowNum,endCol,rowNum)
                        # name cell pattername + device and temp
                        try:
                            name = (row[0].value +str(row[5].value).replace('.','_')\
                                + re.sub('[- ]','',WS.cell(2,column=i).value))
                            workbook.create_named_range(name, WS, colLetter+rowNum)
                        except: 
                            del workbook.defined_names[name]
                            workbook.create_named_range(name, WS, colLetter+rowNum)
            if 'Notes:' in WS.cell(1, i).value : break
            startTempCol = 0

    # get rid of empty columns at the end
    lastCol = WS.max_column
    for i in range(1,WS.max_column+1) :
        if 'Notes:' in str(WS.cell(1, i).value) :
            WS.cell(1, i).fill = orange
            lastCol = i
            break
    WS.merge_cells(start_row=1, start_column=lastCol, end_row=2, end_column=lastCol)
    WS.delete_cols(lastCol+1,WS.max_column)    

    format_worksheet(WS)
    if 'Vm' in WS['I1'].value : WS.column_dimensions['I'].width = 10
    else: WS.column_dimensions['F'].width = 10
    WS.column_dimensions['C'].width = 3.5
    #add red font formatting to bad cells
    for r in range(3,WS.max_row+1):
        for c in range(11,WS.max_column)  :
            if '-' in WS['I1'].value or '-' in WS['F1'].value:
                rule = CellIsRule(operator='greaterThan', stopIfTrue=False,\
                formula=['$I%s'% str(r)], font=redFont)
            else :
                rule = CellIsRule(operator='lessThan', \
                formula=['$I%s'% str(r)], stopIfTrue=False, font=redFont)
            cellStr = get_column_letter(c) + str(r)
            WS.conditional_formatting.add(cellStr, rule)
            failRule = FormulaRule(formula=[('AND(NOT(ISNUMBER(%s)),NOT(ISBLANK(%s)))'%(cellStr,cellStr))]\
                ,font=redFont,fill=PatternFill(start_color='00C0C0C0',end_color='00C0C0C0',fill_type='solid'))
            WS.conditional_formatting.add(cellStr, failRule)
    if last: 
        make_average_tab(workbook,replace)
        for ws in workbook.sheetnames:
            if 'Averages' in ws :
                make_graphs(workbook[ws],workbook)
    

    # put infoTab at the begining
    position = workbook.worksheets.index(workbook['Info'])
    sheets = workbook._sheets.copy()
    sheets.insert(0, sheets.pop(position)) #modifying the sheets list
    workbook._sheets = sheets 
    #organize sheets, averages at beginning
    for sheet in workbook.sheetnames :
        if 'Tm' in workbook[sheet].title :
            workbook[sheet].sheet_properties.tabColor = 'A8D846'
        elif 'Vm' in workbook[sheet].title :
            workbook[sheet].sheet_properties.tabColor ='C6C6C6'
        if 'Graph' in sheet:
            position = workbook.worksheets.index(workbook[sheet])
            sheets = workbook._sheets.copy()
            sheets.insert(1, sheets.pop(position))
            workbook._sheets = sheets
    for sheet in workbook.sheetnames:
        if 'Average' in sheet: 
            position = workbook.worksheets.index(workbook[sheet])
            sheets = workbook._sheets.copy()
            sheets.insert(1, sheets.pop(position))
            workbook._sheets = sheets
    saved = False
    warned = False
    while not saved:
        try: # save the excel workbook
            time.sleep(0.5)
            workbook.save(outputFile)
            if not printed : print('- Done!')   
            else : print ('Done!')
            saved = True
        except KeyboardInterrupt :
            print('\nKeyboard Interrupt: Process Killed')
            sys.exit()
        except Exception as e: # won't save if the file is already open
            if not warned :
                print('\nPlease close the spreadsheet %s first' % outputFile)
                printed = True
                warned = True

def insert_avg(avgSheet, avgData, avgRow, temp, corner) :
    '''Insert the average from one column in min/max sheet into avg sheet'''
    tempCol = 0 # insertion column
    newCol = False
    tempFound = False
    exit = False
    for rows in avgSheet.iter_rows(min_row=1, max_row=1, min_col=6):
        for cell in rows:
            if not isinstance(cell.column, int): # make sure col is number
                    cellColumn = column_index_from_string(cell.column)
            else : cellColumn = cell.column
            if exit :  break
            if cell.value == temp : # if they have the same temperature
                tempFound = True
                cellCorner = avgSheet.cell(2, cellColumn).value
                if len(cellCorner) != 2 : continue
                if corner > cellCorner : # if the corner is greater, insert after
                    tempCol = cellColumn + 1
                    newCol = True
                elif corner == cellCorner : # if same corner, insert in same row
                    tempCol = cellColumn
                    newCol = False
                    exit = True
                else:
                    tempCol = cellColumn
                    newCol = True
                    exit = True
            elif not tempFound : #insert temp before current temp if smaller
                try:
                    if int(cell.value[:cell.value.find('C')]) > \
                        int(temp[:temp.find('C')]) :
                            tempCol = cellColumn
                            newCol = True
                            exit = True
                except : 
                    pass
    # if the temp and device col was not found, add it
    if tempCol == 0 : 
        avgSheet.insert_cols(avgSheet.max_column+1)
        tempCol = avgSheet.max_column+1
        newCol = True
    if newCol:
        avgSheet.insert_cols(tempCol)
        avgSheet.cell(1, tempCol).value = temp # add temp
        avgSheet.cell(2, tempCol).value = corner # add corner
    #insert the data into the new column
    avgSheet.cell(avgRow, tempCol).value = avgData
    
def make_graphs(avgTab,workbook) :
    '''Makes 2D bar graphs for each temp and 3D surface plots for each corner'''
    Tmin = Vmin = False
    if 'Vmin' in avgTab['A1'].value : Vmin = True
    if 'Tmin' in avgTab['A1'].value : Tmin = True
    #add filters to averages tab. must filter by spec voltage to get nice graphs
    avgTab.auto_filter.ref = 'A2:' + get_column_letter(avgTab.max_column)+str(avgTab.max_row)
    avgTab.auto_filter.add_filter_column(5,[])
    avgTab.auto_filter.add_sort_condition('E2:E'+str(avgTab.max_row))
    name = avgTab.title.replace('Averages','Graphs')
    if name in workbook.sheetnames :
        graphTab = workbook[name]
    else : graphTab = workbook.create_sheet(name)
    tempRanges = []
    curInd = 0
    for c in range(5,len(avgTab[1])+1): # get grouping of cols by temp
        if len(tempRanges) == curInd : tempRanges.append([]) # add blank list
        curCell = avgTab.cell(1,c).value
        rightCell = avgTab.cell(1,c+1).value
        if 'C' in curCell :
            if rightCell == None : #last col
                tempRanges[curInd].append(c); break
            elif rightCell != curCell: # last col of temp
                tempRanges[curInd].append(c)
                curInd += 1
            elif curCell == rightCell : # same temp cols next to each other
                tempRanges[curInd].append(c)
    maxValue = 0
    minValue = 100
    for r in range(3,avgTab.max_row) : # get max and min spec voltages for y-axis
        curCell = avgTab.cell(r, 5).value
        if curCell == None : break
        if curCell > maxValue : maxValue = curCell
        if curCell < minValue : minValue = curCell
    for rang in tempRanges :
        legEntry = []
        for column in rang :
            legEntry.append(avgTab.cell(2,column).value)
        namesInChart = 25
        numCharts = max(1,round((avgTab.max_row-2)/namesInChart))
        minRow = 3
        for i in range(1,numCharts+1) : # divide into graphs with around 25 points
            insertRow = str(2 + 25*tempRanges.index(rang))
            insertCol = get_column_letter(2 + 15*(i-1))
            maxRow = min(int(avgTab.max_row/numCharts)*i,avgTab.max_row+1)
            barChart = BarChart()
            barChart.type = 'col'
            barChart.style = 2
            barChart.title = avgTab['A1'].value + ' @ '+ avgTab.cell(1,rang[0]).value
            for col in rang :
                data = Reference(avgTab, min_col=col, min_row=minRow,max_row=maxRow)
                series = Series(data,title=avgTab.cell(2,col).value)
                barChart.append(series)
            categories = Reference(avgTab, min_col=1, min_row=minRow, max_row=maxRow)
            barChart.set_categories(categories)
            barChart.shape = 4
            barChart.height = 12.5 
            barChart.width = 25
            barChart.layout = Layout(manualLayout=ManualLayout(w=0.95,))
            if Vmin : barChart.y_axis.scaling.min = round(0.6 * minValue,2)
            barChart.legend.layout = \
                Layout(manualLayout=ManualLayout(x=0,y=-0.44,h=0.05, w=0.5,))
            lineChart = LineChart()
            specData = Reference(avgTab, min_col=5, min_row=minRow,max_row=maxRow)
            title=''
            if Vmin : title = 'Spec Vmin'
            if Tmin : title = 'Spec Period'
            specSeries = Series(specData, title=title)
            specSeries.graphicalProperties.line.width = 30000
            lineChart.append(specSeries)
            location = insertCol+insertRow
            barChart += lineChart
            graphTab.add_chart(barChart, location)
            minRow = maxRow + 1

    # make 3d graphs 
    maximumRow = avgTab.max_row + 2
    avgTab.insert_rows(maximumRow)  
    if Vmin : avgTab.cell(maximumRow,4).value = 'Spec Period Averages'
    if Tmin : avgTab.cell(maximumRow,4).value = 'Spec Vmin Averages'
    for c in range(6,len(avgTab[1])): #find all tests with same spec period 
        periodAvgs = {}
        for row in avgTab.iter_rows(min_row=3, max_row = maximumRow-2): 
            specP = row[1].value
            if specP == None : continue
            vMin = avgTab.cell(row[0].row,c).value
            if vMin == None: continue
            name = vMin[vMin.find('\''):vMin.find(',')] # get location 
            if specP in periodAvgs.keys(): # add to list
                    vList= periodAvgs[specP][0]
                    vList += ',IF(ISNUMBER('+ name +'),'+ name +' ,0)'
                    periodAvgs[specP][0] = vList
                    cList= periodAvgs[specP][1]
                    cList += '+IF(ISNUMBER('+ name +'),1,0)'
                    periodAvgs[specP][1] = cList
            else : 
                insertTuple = ['IF(ISNUMBER('+ name +'),'+ name +' ,0)','IF(ISNUMBER('+ name +'),1,0)']
                periodAvgs[specP] =  insertTuple
        if avgTab.max_row == maximumRow +1: avgTab.insert_rows(maximumRow+1,len(periodAvgs))
        for period in periodAvgs.keys(): # make excel formula to find average of period
            for row in avgTab.iter_rows(min_row = maximumRow+1,max_row= maximumRow+len(periodAvgs)): 
                list = periodAvgs[period]
                avgFormula = '=ROUND(SUM('+list[0]+')/SUM('+list[1]+'),2)'
                if period == row[4].value :
                    curRow = row[0].row
                    avgTab.cell(curRow,c).value = avgFormula
                    break
                elif row[4].value == None :
                    row[4].value = period
                    curRow = row[0].row
                    avgTab.cell(curRow,c).value = avgFormula
                    break
                elif period < row[4].value :
                    curRow = row[0].row
                    avgTab.insert_rows(curRow)
                    avgTab.cell(curRow,5).value = period
                    avgTab.cell(curRow,c).value = avgFormula
                    break
    cornerDict = {}
    for c in range(6,len(avgTab[1])+1): # get grouping of cols by corner
        corner = avgTab.cell(2,c).value
        if corner != None:
            if corner in cornerDict.keys(): 
                colList = cornerDict[corner]
                colList.append(c)
                cornerDict[corner] = colList
            else : cornerDict[corner] = [c]
    for key in cornerDict.keys(): # make surface plot for each corner
        surfChart = SurfaceChart()
        for col in cornerDict[key] :
                data = Reference(avgTab, min_col=col, min_row=maximumRow+1,max_row=avgTab.max_row)
                series = Series(data,title=avgTab.cell(1,col).value)
                surfChart.append(series)
        labels = Reference(avgTab, min_col=5, min_row=maximumRow+1, max_row=avgTab.max_row)
        surfChart.set_categories(labels)
        surfTitle=''
        if Vmin : surfTitle = key+' Spec Period vs. Vmin vs. Temp'
        if Tmin : surfTitle = key+' Spec Vmin vs. Tmin vs. Temp'
        surfChart.title = surfTitle
        if Vmin :
            surfChart.y_axis.title = 'Vmin (V)'
            surfChart.x_axis.title = 'Spec Period (ns)'
        if Tmin:
            surfChart.x_axis.title = 'Spec Vmin (V)'
            surfChart.y_axis.title = 'Tmin (ns)'
        if Vmin : surfChart.y_axis.scaling.min = round(0.6 * minValue,1)
        surfChart.height = 12.5 ; surfChart.width = 25 
        # this is needed for 3d to work on linux for some reason
        surfChart.view3D = None; surfChart.floor = None
        surfChart.sideWall = None; surfChart.backWall = None
        surfChart.y_axis.delete = False
        insertRow = str(int(insertRow)+25) # 25 rows down
        location = 'B'+insertRow
        graphTab.add_chart(surfChart, location)
        
def make_average_tab(workbook,replace) :
    '''Pulls all averages from main sheets and puts them into one tab'''
    dataSheets = []
    for sheet in workbook.sheetnames :
        if ('min' in sheet or 'max' in sheet) and not('Averages' in sheet) and\
            not('Graphs' in sheet):
            dataSheets.append(sheet)
    if replace:
        for sheet in workbook.sheetnames :
            if 'Averages' in sheet : del workbook[sheet]
    for sheet in dataSheets :  
        WS = workbook[sheet]
        spec = sheet[:sheet.find('_')] # e.g Vmin(-5%)
        if (spec + '_Averages') in workbook.sheetnames:
            avgTab = workbook[spec + '_Averages']
        else: avgTab = workbook.create_sheet(spec + '_Averages')
        if avgTab['A1'].value == None :
            # write in the basic headers
            if 'v' in spec.lower() : avgTab['A1'] = 'Vmin at Spec Period (V)'
            elif 't' in spec.lower() : avgTab['A1'] = 'Tmin at Spec Vmin (ns)'
            avgTab['A2'] = 'Shmoo Tests' 
            for r in range(1, 3):
                for c in range(2,6):
                    if avgTab.cell(r,c).coordinate in avgTab.merged_cells: continue
                    avgTab.cell(r,column=c).value = \
                        workbook[sheet].cell(r,column=c+4).value
            avgTab.merge_cells('C1:D1')
            avgTab.row_dimensions[1].height = 30

        avg_columns = []
        for col in range(1,len(WS[1])+1) :
            if WS.cell(1,col).value == 'Average' :
                avg_columns.append(col)
        for col in avg_columns :
            temp = WS.cell(1,col-1).value.replace(' ','')
            corner = re.sub('[0-9\-C ]', '', WS.cell(2,col).value)
            for dataRow in WS.iter_rows(min_row=3) : # iterate over data in column
                patternName = dataRow[0].value
                patternName =  patternName.replace('_Shmoo','').replace('_shmoo','')
                patternName =  patternName.replace('_shm','').replace('_Shm','')
                cellName = dataRow[0].value +str(float(dataRow[5].value)).replace('.','_') + temp.replace('-','') + corner
                avgData = '=IFERROR(' + workbook.defined_names.get(cellName).attr_text +',"")'
                patternFound = False
                for avgRow in avgTab.iter_rows(min_row=3) : # iter through avg sheet
                    # check if name and same test params already written
                    if avgRow[0].value == patternName and \
                        avgRow[1].value == dataRow[5].value and \
                        avgRow[2].value == dataRow[6].value and \
                        avgRow[3].value == dataRow[7].value and \
                        avgRow[4].value == dataRow[8].value :
                            patternFound = True
                            insert_avg(avgTab, avgData, avgRow[0].row, temp, corner)
                            break
                if not patternFound : # append if not found 
                    avgTab.append([patternName,dataRow[5].value,dataRow[6].value,\
                        dataRow[7].value, dataRow[8].value])
                    insert_avg(avgTab, avgData, avgTab.max_row, temp, corner)
        format_worksheet(avgTab)
    for sheet in workbook.sheetnames :
        if not 'Averages' in sheet : continue
        avgTab = workbook[sheet]
        if 'V' in avgTab['E2'].value or 'V' in avgTab.title: 
            avgTab.column_dimensions['E'].width = 10
        else : avgTab.column_dimensions['B'].width = 10
        #add red font formatting to bad cells
        for r in range(3,avgTab.max_row+1):
            for c in range(6,avgTab.max_column+1)  :
                if '-' in avgTab['E1'].value or '-' in avgTab['B1'].value:
                    rule = CellIsRule(operator='greaterThan', stopIfTrue=False,\
                    formula=['$E%s'% str(r)], font=redFont)
                else :
                    rule = CellIsRule(operator='lessThan', \
                    formula=['$E%s'% str(r)], stopIfTrue=False, font=redFont)
                cellStr = get_column_letter(c) + str(r)
                avgTab.conditional_formatting.add(cellStr, rule)
                failRule = FormulaRule(font=redFont, formula=\
                    [('AND(NOT(ISNUMBER(%s)),NOT(ISBLANK(%s)))'%(cellStr,cellStr))])
                avgTab.conditional_formatting.add(cellStr, failRule)


def make_info_tab(workbook, csv, last, replace, new):
    '''Creates tab with data of where tests files are and how this report was made'''
    try :
        infoTab = workbook['Info']
    except:
        infoTab = workbook.create_sheet('Info')
    #where the openfive logo is located
    infoTab._images = []
    imgDir = os.path.abspath(os.path.join(os.path.realpath(__file__), os.pardir))
    imgPath = os.path.join(imgDir,'openfive_logo.png')
    img = Image.open(imgPath)
    img = img.resize((390,74),Image.NEAREST)
    img.save(imgPath)
    img = openpyxl.drawing.image.Image(imgPath)
    img.anchor = 'G2'
    infoTab.add_image(img)
    infoTab['H6'] = 'Script:'
    infoTab['H7'] = 'Version:' 
    infoTab['H8'] = 'Creation:' 
    infoTab['I6'] = ' ' + os.path.basename(sys.argv[0])
    version = get_version(sys.argv[0])
    infoTab['I7'] = ' rev ' + version
    infoTab['I8'] = ' ' + datetime.now().strftime('%#I:%M%p, %m/%d/%Y ')
    # align info header labels
    for row in infoTab.rows :
        if row[0].row > 1 :
            row[7].alignment = Alignment(horizontal='right',wrap_text=False)
    # add legend
    infoTab['A2'].value = 'LEGEND:'
    infoTab['A2'].font = Font(size = 14, bold = True)
    infoTab['A3'].value = 'FAIL'
    infoTab['A3'].font = redFont
    infoTab['A3'].border = thinBorders
    infoTab['A3'].alignment = Alignment(horizontal='center')
    infoTab['A3'].fill = grey # grey
    infoTab['B3'] = 'There is no voltage where test passes at spec period' 
    infoTab['A4'].fill = blue
    infoTab['A4'].border = thinBorders
    infoTab['B4'] = 'There exist significant holes in the shmoo plot' 
    infoTab['A5'].fill = green
    infoTab['A5'].border = thinBorders
    infoTab['B5'] = 'The spec period is not in passing range @ spec Voltage' 
    infoTab['A6'].value = 'Num'
    infoTab['A6'].alignment = Alignment(horizontal='center')
    infoTab['A6'].font = redFont # red font if not passing
    infoTab['A6'].border = thinBorders
    infoTab['B6'] = 'The observed min/max does not meet spec'
    # add all paths to list
    infoTab['A9'] = 'Paths to Data Files:'
    infoTab['A9'].font = Font(bold=True)
    lastRow = 9
    for row in infoTab.rows :
        if row[0].row > 9 : # find the last row containing a path
            if row[0].value or row[1].value : lastRow = row[0].row
    infoTab.cell(lastRow+1, 1).value = None # add blank line
    infoTab.cell(lastRow+2, 1).value = os.path.abspath(csv) #add csv title
    infoTab.cell(lastRow+2, 1).font = Font(bold=True,size=14) 
    infoTab.cell(lastRow+2,3).value = None
    lastRow += 2
    topRow = lastRow + 1
    patternName = ''
    with open(csv) as csvFile :
        lines = csvFile.readlines()
        for line in lines :
            if 'Pattern Name, EQNSET #, Holes, Bad' in line: continue
            data = line.split(',')
            filePath = data[14]
            test = data[9] + '-' + data[10].replace(' ','')
            if patternName != data[0] : #insert new pattern name under csv
                infoTab.insert_rows(lastRow+1,1)
                infoTab.cell(lastRow+1,1).value = data[0]
                lastRow += 1
                patternName = data[0]
                infoTab.insert_rows(lastRow+1,2)
                lastRow +=2 
            column = 2
            for c in range(2,infoTab.max_column+1):
                if infoTab.cell(lastRow,c).value != ' ' and \
                infoTab.cell(lastRow,c).value != None: column +=1
                else : break
            infoTab.column_dimensions[get_column_letter(column)].width = 10
            infoTab.cell(lastRow-1,column).value = test
            infoTab.cell(lastRow-1,column).font = Font(size=12)
            infoTab.cell(lastRow,column).value = \
                os.path.abspath(os.path.normpath(filePath.replace(' ','')))
            infoTab.cell(lastRow,column).alignment = \
                Alignment(horizontal='left',vertical='top')
            infoTab.cell(lastRow,column+1).value = ' ' # stop last file from overflow

    #collapse list of files so it's readable
    infoTab.insert_rows(lastRow+2,1) #add row at bottom 
    infoTab.cell(lastRow+2,3).value = ' ' # new row needs value
    infoTab.sheet_properties.outlinePr.summaryBelow = False #put expand button at top
    infoTab.row_dimensions.group(topRow, lastRow, outline_level=1, hidden=True)
    firstRow = 0
    infoTab.cell(1,1).value = None
    #make subgroups of file names
    for row in infoTab.rows :
        if row[0].row > 9 :
            if firstRow == 0 and row[1].value : firstRow = row[0].row # new pattern
            elif firstRow > 0 and row[1].value : continue # same pattern
            elif firstRow > 0 and row[1].value == None : #end of pattern
                infoTab.row_dimensions.group(firstRow, row[0].row-1, \
                    outline_level=2,hidden=True)
                firstRow = 0
            else : firstRow = 0

def format_worksheet(WS) :
    ''' Takes the data worksheet and formats the column widths and cell formats'''
    # format the column width
    dims = {}
    notesCol = WS.max_column
    for row in WS.rows:
        for cell in row:
            if not isinstance(cell.column, int) :
                cellColumn = column_index_from_string(cell.column)
                columnLetter = cell.column
            else :
                cellColumn = cell.column
                columnLetter = get_column_letter(cell.column)
            if 'Notes:' in str(cell.value) : notesCol = cellColumn
            if cell.coordinate in WS.merged_cells: # not merged cells
                continue
            elif cell.value == None and cell.row > 1: # empty columns get smaller
                dims[columnLetter] = max((dims.get(columnLetter, 0), 3))
            elif cell.value and cell.row > 1 : # get max size cell in column
                dims[columnLetter] = \
                    max( (dims.get(columnLetter, 0), len(str(cell.value))) )
    for col, value in dims.items(): # set the value for each column
        if col == 'A' : WS.column_dimensions[col].width = value 
        if 'Averages' in WS.title and column_index_from_string(col) > 5 : 
            WS.column_dimensions[col].width = 8
        elif WS[col+'1'].value == 'Average' : WS.column_dimensions[col].width = 8
        elif col == 'G' or col == 'H' : WS.column_dimensions[col].width = value * 1.5
        else: WS.column_dimensions[col].width = value * 1.3

    # set the formatting for all cells 
    alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=False)
    for r in range(1,WS.max_row+1) :
        for c in range(1,WS.max_column+1) :
            curCell = WS.cell(r, c)
            # border format for all cells
            curCell.border = thinBorders
            if r==2 :
                curCell.border = Border(left=curCell.border.left,\
                    right=curCell.border.right, top=curCell.border.top,\
                      bottom=Side(style='thick'))
            if '(V)' in str(WS.cell(2,column=c).value) and 'Vm' in WS.title :
                curCell.border = Border(left=curCell.border.left,\
                 right=Side(style='thick'), top=curCell.border.top,\
                      bottom=curCell.border.bottom)
            if '(ns)' in str(WS.cell(2,column=c).value) and 'Tm' in WS.title :
                curCell.border = Border(left=curCell.border.left,\
                 right=Side(style='thick'), top=curCell.border.top,\
                      bottom=curCell.border.bottom)
            if r < 3 : 
                 curCell.font = Font(bold=True)
                 if 'Averages' in WS.title:
                    curCell.fill = grey
            if c == notesCol and not 'Averages' in WS.title: # notes formatting
                curCell.alignment = \
                    Alignment(horizontal='left',vertical='bottom',wrap_text=False)
            elif r == 2 : 
                curCell.alignment = \
                    Alignment(horizontal='center',vertical='center',wrap_text=False)
            elif c == 1 and r > 2:
                curCell.alignment = \
                    Alignment(horizontal='left',vertical='bottom',wrap_text=False)
            elif r == 1 : # first row : pattern format
                WS.cell(1, c).alignment = \
                    Alignment(horizontal='center',vertical='center',wrap_text=True)
            else : # all other rows, match fill with corresponding pattern name cell
                curCell.alignment = alignment
                if curCell.fill == PatternFill(None) and c < 10: 
                    curCell.fill = copy(WS.cell(r,column=1).fill)

def csv_to_worksheet(wSheet, csvFile, commenting) :
    ''' Writes the given CSV file into the worksheet in excel
    Input:
        csvFile     --> given .csv file to look in
        wSheet      --> given worksheet to work with
    Output:
        Written CSV data in an excel format on spreadsheet
    '''
    with open(csvFile) as csv :
        lines = csv.readlines()
        for line in lines :
            data = line.split(', ') # get list of data from csv line    
            if 'Pattern Name' in line :
                if 'Vm' in data[8] :
                    wSheet['G1'].value = 'Test Range - VDDcore'; wSheet['I2'].value = '(V)'
                    wSheet['F1'].value = 'Spec Period'; wSheet['F2'].value = '(ns)'
                    wSheet['I1'].value = 'Spec      ' + data[8].replace('Spec','')
                else:
                    wSheet['G1'].value = 'Test Range - Period'; wSheet['I2'].value = '(ns)'
                    wSheet['I1'].value = 'Spec Period'; wSheet['F2'].value ='(V)'
                    wSheet['F1'].value = 'Spec      ' + data[5].replace('Spec','')
                wSheet.merge_cells('G1:H1')
                wSheet.cell(1, wSheet.max_column+1).value = 'Notes: '
                continue
            data[5] = data[5].replace('ns','')
            data[8] = data[8].replace('ns','')
            for i in range(1,len(data)) :
                try:
                    data[i] = round(float(data[i].strip(' ')),3) # read numbers 
                except: continue
            patternFound = False
            rowNum = 3
            for row in wSheet.iter_rows(min_row=3) : 
                # check if name and same test params already written
                if row[0].value == data[0] and row[1].value == data[1] and \
                    row[5].value == data[5] and row[6].value == data[6] and \
                        row[7].value == data[7] and row[8].value == data[8]:
                    patternFound = True
                    insert_data(wSheet, rowNum, data,commenting)
                rowNum += 1
            if not patternFound : # append if not found 
                wSheet.append(data[0:9])
                insert_data(wSheet, wSheet.max_row, data,commenting)
    return wSheet

def insert_data(wSheet, rowNum, data, commenting) :
    ''' Writes one data point to the excel worksheet
    Input:
        wSheet      --> given worksheet to work with
        rowNum      --> row that is being worked on in worksheet
        data        --> data point from one .csv file line
    Output:
        Written data point to given worksheet
    '''
    #check the first row for the temp 
    tempCol = 0 # insertion column
    deviceNum = int(re.sub('[^0-9]','', data[10]))
    newCol = False
    tempFound = False
    exit = False
    for rows in wSheet.iter_rows(min_row=1, max_row=1, min_col=10):
        for cell in rows:
            if not isinstance(cell.column, int):
                    cellColumn = column_index_from_string(cell.column)
            else : cellColumn = cell.column
            if exit :  break
            if cell.value == data[9] : # if they have the same temp
                tempFound = True
                dNum = int(re.sub('[^0-9]', '', wSheet.cell(2, cellColumn).value))
                if deviceNum > dNum : # if the device number is greater, insert after
                    tempCol = cellColumn + 1
                    newCol = True
                elif len(str(dNum)) > 0 and deviceNum == dNum : # if same device, insert in same row
                    tempCol = cellColumn
                    newCol = False
                    exit = True
                else:
                    tempCol = cellColumn
                    newCol = True
                    exit = True
            elif not tempFound : #insert temp before current temp if smaller
                try:
                    if int(cell.value[:cell.value.find('C')]) > \
                        int(data[9][:data[9].find('C')]) :
                            tempCol = cellColumn
                            wSheet.insert_cols(tempCol)
                            newCol = True
                            exit = True
                except : 
                    pass
    # if the temp and device col was not found, add it
    if tempCol == 0 : 
        wSheet.insert_cols(wSheet.max_column)
        tempCol = wSheet.max_column
        newCol = True
    if newCol:
        wSheet.insert_cols(tempCol)
        wSheet.cell(1, tempCol).value = data[9] # add temp
        wSheet.cell(1, tempCol).fill = highlight
        wSheet.cell(2, tempCol).value = data[10] # add device name
    #insert the data into the new column
    insertCell = wSheet.cell(rowNum, tempCol)
    nameCell = wSheet.cell(rowNum,column=1)
    comment = Comment(os.path.abspath(os.path.normpath(data[14])), __file__)
    comments = False
    if 'All Failed' in data[12] or 'All periods fail' in data[12]:
        insertCell.value = 'FAIL' # highlight grey if failed
        # insertCell.fill = grey
        nameCell.fill = grey
        comments = True
    else : 
        insertCell.value = float(data[11])
        notesValue = str(wSheet.cell(rowNum, wSheet.max_column).value)
        # not passing min
        if ('-' in wSheet['I1'].value or '-' in wSheet['F1'].value) and \
            (data[11] > float(wSheet.cell(rowNum, 9).value)) :
            #insertCell.font = redFont # red font if not passing
            if 'Vmin' in wSheet.title :
                wSheet.cell(rowNum, 4).value = 'Vmin'
            if 'Tmin' in wSheet.title :
                wSheet.cell(rowNum, 4).value = 'Tmin'
            comments = True
        # not passing max
        elif not ('-' in wSheet['I1'].value or '-' in wSheet['F1'].value) and \
            data[11] < float(wSheet.cell(rowNum, 9).value) :
            #insertCell.font = redFont # red font if not passing
            if 'Vmax' in wSheet.title :
                wSheet.cell(rowNum, 4).value = 'Vmax'
            if 'Tmax' in wSheet.title :
                wSheet.cell(rowNum, 4).value = 'Tmax'
            comments = True
        else:
            insertCell.fill = PatternFill(None)
            insertCell.font = Font(None)
        # include test in notes if there are holes
        if 'Holes' in data[12] or 'Failing at' in data[12]:
            test = '(' + str(data[9]) +': ' + str(data[10]) + '), '
            tempStartIndex = notesValue.find(('('+str(data[9])+':'))+1 #temp index
            tempChunk = notesValue[tempStartIndex:notesValue[tempStartIndex:].find(')')+\
                    tempStartIndex] 
            if notesValue == 'None' :
                notesValue = 'With Holes:' + test
            elif not ('With Holes:' in notesValue) :
                notesValue = notesValue + 'With Holes: ' + test
            elif tempStartIndex < 1 :
                notesValue = notesValue + test
            elif (tempStartIndex != -1) and not str(data[10]) in tempChunk:
                notesValue = (notesValue[:tempStartIndex] + tempChunk + \
                    ', ' + str(data[10]) + notesValue[tempStartIndex + \
                    notesValue[tempStartIndex:].find(')'):])
            wSheet.cell(rowNum, 3).value = 'VH'
            if insertCell.fill == PatternFill(None) :
                insertCell.fill = blue 
                if nameCell.fill != grey :
                    nameCell.fill = blue
            comments = True
        # if there are issues in shmoo at spec Vmin
        if str(data[8]) in data[12] or 'Spec period does not' in data[12] : 
            # undetermined = 'Undeterminable at %sV; ' % str(data[8])
            # if 'Spec period not' in data[12] :
            #     undetermined = 'Spec period not in passing range; '
            # if notesValue == 'None' :
            #     notesValue = undetermined
            # elif not (undetermined in notesValue) :
            #     notesValue = undetermined + notesValue
            if 'Vm' in wSheet['I1'].value :
                wSheet.cell(rowNum, 5).value = data[8]
            else : wSheet.cell(rowNum, 5).value = data[5]
            insertCell.fill = green 
            if nameCell.fill != grey : nameCell.fill = green
            comments = True
        if 'Period on edge' in data[12] : # spec period is on edge of passing periods
            if notesValue == 'None' :
                notesValue = data[12]
            elif not ('Period on edge' in notesValue) :
                notesValue = 'Period on edge of passing range; ' + notesValue
            comments = True
        wSheet.cell(rowNum,wSheet.max_column).value = notesValue.strip('None')
    if commenting and comments : insertCell.comment = comment
    

if __name__=='__main__' :
    parser = argparse.ArgumentParser(description='Transform CSV into Excel',\
        formatter_class=argparse.RawTextHelpFormatter,\
        epilog = 'usage examples:\n'\
        '  shmoo_csv_to_excel -i Vmin_-5p_TT.csv -o excelfiles/TT/ -n\n\n'\
        '  shmoo_csv_to_excel -i Vmin_-5p_TT.csv -f Vmin_TT.xlsx -s\n')
    parser.add_argument('-v', '-V', '--version', dest='version', action='store_true',\
        default=False, help='get version of script and exit')
    parser.add_argument('-i', '--input', nargs='+', dest='inputCSV', default=None, \
        help='required input CSV location')
    parser.add_argument('-f', '--filename', dest='outputFile', default=None, \
        help='output file name.')
    parser.add_argument('-o', '--output', dest='outputDir', default='.', \
        help='output folder name/location. DEFAULT current folder.')
    parser.add_argument('-s', '--replace', dest='replace', action='store_true', \
        default=False, help='write over existing excel instead of making new file')
    parser.add_argument('-n', '--notes', dest='comment', action='store_true', \
        default=False, help='add path in an excel note for all problematic tests') 
    args = parser.parse_args()
    if args.version :
        version = get_version(sys.argv[0])
        print('Version: ' + version)
        sys.exit()
    if args.inputCSV == None :
        parser.print_usage()
        print('Input file is required. Please include using -i')
        sys.exit()
    for csv in args.inputCSV :
        try: 
            shmoo_csv_to_excel(csv,args.outputFile,args.outputDir,args.replace,\
            False,args.comment,True)
        except KeyboardInterrupt :
            print('\nKeyboard Interrupt: Process Killed')
            sys.exit()